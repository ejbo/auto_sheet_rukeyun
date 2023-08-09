import datetime

from openpyxl import Workbook, load_workbook, utils


room_name_ws1 = ["福2", "福3", "555", "666", "777", "888", "999", "101", "102", "110", "111", "112", "113", "115", "116", "117", "118", "外卖"]
room_name_ws2 = ["V1包", "V2包", "V3包", "V5包", "V6包", "01台", "02台", "03台", "05台", "06台", "07台", "08台", "09台", "10台", "外卖"]
room_name_ws3 = ["111包", "222包", "333包", "11台", "12台", "13台", "15台", "16台", "17台", "18台", "19台", "20台", "外卖"]


def create_workbook(excel_name, sheet1, sheet2, sheet3):
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet(sheet1)
    ws2 = wb.create_sheet(sheet2)
    ws3 = wb.create_sheet(sheet3) 
    
    # Title
    ws1.merge_cells('A1:CL1')
    ws2.merge_cells('A1:BW1')
    ws3.merge_cells('A1:BM1')
    
    current_time = datetime.datetime.now().strftime('%Y 年 %m 月 %d 日 %H 时 %M 分 %S 秒')
    ws1['A1'] = "海 鱼 馆 销 售 时 录 明 细 表" + " （ " + current_time + " ）"
    ws2['A1'] = "老 菜 馆 销 售 时 录 明 细 表" + " （ " + current_time + " ）"
    ws3['A1'] = "水 煮 鱼 销 售 时 录 明 细 表" + " （ " + current_time + " ）"

    # Title Style (Unifinished)

    generate_first_row_para(ws1, "CL", room_name_ws1)
    generate_first_row_para(ws2, "BW", room_name_ws2)
    generate_first_row_para(ws3, "BM", room_name_ws3)

    generate_content_cell(ws1, "CL")
    generate_content_cell(ws2, "BW")
    generate_content_cell(ws3, "BM")

    print(3)
    
    wb.save(excel_name)

def generate_first_row_para(worksheet, end_column_string, room_name_list):

    # Get the maximum column index (e.g., column 'CL' corresponds to index 90)
    max_col_index = utils.column_index_from_string(end_column_string)

    # Specify the start row
    start_row = 2

    # Merge every three cells after two cells
    col = 1  # Start from column C

    room_name_list_index = 0

    while col <= max_col_index:

        worksheet.cell(row = start_row, column = col, value = "下单时间")
        col += 1
        
        worksheet.cell(row = start_row, column = col, value = "结账时间")
        col+= 1

        start_cell = worksheet.cell(row=start_row, column=col)
        end_col = min(col + 2, max_col_index)  # Ensure we don't exceed the max column
        end_cell = worksheet.cell(row=start_row, column=end_col)
        worksheet.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
        
        worksheet.cell(row = start_row, column = col, value = room_name_list[room_name_list_index])
        room_name_list_index += 1
        col = end_col + 1  # Move to the next set of cells

def generate_content_cell(worksheet, end_column_string):

    max_col_index = utils.column_index_from_string(end_column_string)

    for col in range(1, max_col_index + 1):

        if (col % 5 == 0):
            for row in range (3, 50, 6):
                start_cell = worksheet.cell(row=row, column=col)
                end_cell = worksheet.cell(row=row + 5, column=col)
                worksheet.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
            continue

        for row in range(3, 50, 2):
            start_cell = worksheet.cell(row=row, column=col)
            end_cell = worksheet.cell(row=row + 1, column=col)
            worksheet.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)

    for col in range(3, max_col_index + 1, 5):
        for row in range(3, 50, 6):
            worksheet.cell(row = row, column = col, value = "海鱼馆")
            worksheet.cell(row = row + 2, column = col, value = "老菜馆")
            worksheet.cell(row = row + 4, column = col, value = "水煮鱼")

create_workbook("订单统计表.xlsx", "海鱼馆", "老菜馆", "水煮鱼")
# initiate_sheet("订单统计表.xlsx", "小资味海鱼馆")
